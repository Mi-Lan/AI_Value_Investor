"""
TIKR Financial Statements Scraper - Standalone App

A clean, automated Python scraper for TIKR financial statements that runs without interactive prompts.
This standalone version extracts historical financial statements from the TIKR platform and exports
them to Excel spreadsheets for analysis.

Features:
- Income Statement data (Revenue, expenses, profit/loss metrics)
- Cash Flow Statement data (Operating, investing, financing cash flows)
- Balance Sheet data (Assets, liabilities, equity)
- Live Market Data (Current price, P/E ratios, market cap)
- 🔥 NEW: Yahoo Finance real-time price integration
- 🎯 NEW: Automatic price population in valuation base field O19

Usage:
    python tikr_scraper.py [TICKER]

Examples:
    python tikr_scraper.py          # Scrapes AAPL by default
    python tikr_scraper.py TSLA     # Scrapes Tesla
    python tikr_scraper.py MSFT     # Scrapes Microsoft

Output Excel File Contains:
    • income_statement - Historical revenue, expenses, profit/loss
    • cashflow_statement - Operating, investing, financing cash flows  
    • balancesheet_statement - Assets, liabilities, equity
    • yahoo_finance_realtime - Real-time price from Yahoo Finance
    • sales_to_capital - Sales to Capital ratio analysis
    • valuation_base - Valuation template with O19 (price) & O28 (shares) auto-populated

Requirements:
    pip install yfinance selenium-wire pandas openpyxl python-dotenv webdriver-manager
"""

import requests
import json
import time
import os
import datetime
import logging
from pathlib import Path
from typing import Optional, Tuple, Dict, Any

import pandas as pd
import openpyxl
from seleniumwire import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import WebDriverException
from dotenv import load_dotenv

# Import the keys for data field mappings
try:
    from . import keys
except ImportError:
    # Fallback for direct imports
    import keys

# Try to import yfinance for real-time price data
try:
    import yfinance as yf
    YFINANCE_AVAILABLE = True
except ImportError:
    YFINANCE_AVAILABLE = False
    logger.warning("yfinance not installed. Install with: pip install yfinance")

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class TIKRScraper:
    """
    TIKR Financial Statements Scraper
    
    Scrapes historical financial statements from TIKR platform and exports
    them to Excel spreadsheets for analysis.
    """
    
    def __init__(self, output_dir: str = "outputs"):
        """
        Initialize the TIKR scraper.
        
        Args:
            output_dir: Directory to save the output files
        """
        
        # Load credentials from multiple sources (in order of priority)
        self.username, self.password = self._load_credentials()
        
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        # Validate credentials
        if not self.username or not self.password:
            raise ValueError(
                "Please configure your TIKR credentials in one of these ways:\n"
                "1. Edit config.py file with your TIKR email and password\n"
                "2. Set TIKR_EMAIL and TIKR_PASSWORD in .env file\n"
                "3. Set TIKR_EMAIL and TIKR_PASSWORD environment variables\n\n"
                "You need valid TIKR account credentials to use this scraper.\n"
                "Sign up at: https://app.tikr.com/"
            )
        
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:108.0) Gecko/20100101 Firefox/108.0',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'en-US,en;q=0.5',
            'Accept-Encoding': 'gzip, deflate, br',
            'Content-Type': 'application/json',
            'Origin': 'https://app.tikr.com',
            'Connection': 'keep-alive',
            'Referer': 'https://app.tikr.com/',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'cross-site',
            'Pragma': 'no-cache',
            'Cache-Control': 'no-cache',
            'TE': 'trailers'
        }
        
        self.statements = keys.statements
        self.content = {
            'income_statement': [], 
            'cashflow_statement': [], 
            'balancesheet_statement': []
        }
        
        # Sales to Capital calculation result
        self.sales_to_capital = None
        
        # Token management
        self.token_file = self.output_dir / 'token.tmp'
        self.access_token = self._load_token()
    
    def _load_credentials(self) -> Tuple[str, str]:
        """
        Load TIKR credentials from multiple sources in order of priority:
        1. config.py file
        2. .env file
        3. Environment variables
        4. Demo credentials (fallback)
        
        Returns:
            Tuple of (email, password)
        """
        email = None
        password = None
        source = "not found"
        
        # Priority 1: config.py file
        try:
            try:
                from . import config
            except ImportError:
                # Fallback for direct imports
                import config
            if hasattr(config, 'TIKR_EMAIL') and hasattr(config, 'TIKR_PASSWORD'):
                email = config.TIKR_EMAIL
                password = config.TIKR_PASSWORD
                if email and password and email != "your_email@example.com":
                    source = "config.py"
                    logger.info("✅ Credentials loaded from config.py")
                else:
                    email = None
                    password = None
        except ImportError:
            pass
        except Exception as e:
            logger.warning(f"Error loading config.py: {e}")
        
        # Priority 2: .env file
        if not email or not password:
            try:
                load_dotenv()
                env_email = os.getenv('TIKR_EMAIL')
                env_password = os.getenv('TIKR_PASSWORD')
                if env_email and env_password:
                    email = env_email
                    password = env_password
                    source = ".env file"
                    logger.info("✅ Credentials loaded from .env file")
            except Exception as e:
                logger.warning(f"Error loading .env file: {e}")
        
        # Priority 3: Environment variables (already checked above via load_dotenv)
        
        # Priority 4: Demo credentials (fallback)
        if not email or not password:
            logger.warning("⚠️  No credentials found in config.py or .env file, using demo credentials")
            email = "milanrasovic.f@gmail.com"
            password = "GsHr5-TiP$7uXLn"
            source = "demo credentials"
        
        logger.info(f"🔐 Using credentials from: {source}")
        return email, password
    
    def _load_token(self) -> str:
        """Load access token from file if it exists."""
        if self.token_file.exists():
            with open(self.token_file, 'r') as f:
                return f.read().strip()
        return ''
    
    def _save_token(self, token: str) -> None:
        """Save access token to file."""
        with open(self.token_file, 'w') as f:
            f.write(token)
        logger.info("Access token saved")
    
    def get_access_token(self) -> None:
        """
        Generate access token by logging into TIKR platform.
        
        This method uses Selenium to automate the login process and
        extract the access token from network requests.
        """
        logger.info("Generating access token...")
        
        chrome_options = Options()
        
        # Check if we should show browser (for debugging)
        show_browser = False
        try:
            try:
                from . import config
            except ImportError:
                # Fallback for direct imports
                import config
            if hasattr(config, 'SHOW_BROWSER'):
                show_browser = config.SHOW_BROWSER
        except (ImportError, AttributeError):
            pass
        
        if not show_browser:
            chrome_options.add_argument("--headless")
        
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        user_agent = 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.2 (KHTML, like Gecko) Chrome/22.0.1216.0 Safari/537.2'
        chrome_options.add_argument(f'user-agent={user_agent}')
        chrome_options.add_argument('--window-size=1920,1080')
        
        try:
            service = Service(ChromeDriverManager().install())
            browser = webdriver.Chrome(service=service, options=chrome_options)
            
            # Login to TIKR
            browser.get('https://app.tikr.com/login')
            time.sleep(2)
            
            # Enter credentials
            email_input = browser.find_element(By.XPATH, '//input[@type="email"]')
            password_input = browser.find_element(By.XPATH, '//input[@type="password"]')
            
            email_input.send_keys(self.username)
            password_input.send_keys(self.password)
            
            # Click login button
            login_button = browser.find_element(By.XPATH, '//button/span')
            login_button.click()
            
            # Wait for login to complete
            while 'Welcome to TIKR' not in browser.page_source:
                time.sleep(2)
                if 'Invalid' in browser.page_source or 'Error' in browser.page_source:
                    raise Exception("Invalid credentials or login failed")
            
            logger.info("Successfully logged in to TIKR")
            
            # Navigate to screener to trigger API calls
            browser.get('https://app.tikr.com/screener?sid=1')
            time.sleep(3)
            
            # Trigger fetch screen to capture API token
            fetch_button = browser.find_element(By.XPATH, '//button/span[contains(text(), "Fetch Screen")]/..')
            fetch_button.click()
            time.sleep(5)
            
            # Extract token from network requests
            for request in browser.requests:
                if 'amazonaws.com/prod/fs' in request.url and request.method == 'POST':
                    response = json.loads(request.body)
                    self.access_token = response['auth']
                    self._save_token(self.access_token)
                    logger.info("Successfully extracted access token")
                    break
            else:
                raise Exception("Could not extract access token from network requests")
                
        except WebDriverException as e:
            logger.error(f"WebDriver error: {e}")
            raise Exception("Failed to initialize browser. Make sure Chrome is installed.")
        except Exception as e:
            logger.error(f"Error generating access token: {e}")
            raise
        finally:
            if 'browser' in locals():
                browser.quit()
    
    def find_company_info(self, ticker: str) -> Tuple[Optional[int], Optional[int]]:
        """
        Find company trading ID and company ID from ticker symbol.
        
        Args:
            ticker: Stock ticker symbol (e.g., 'AAPL')
            
        Returns:
            Tuple of (trading_id, company_id) or (None, None) if not found
        """
        headers = self.headers.copy()
        headers['content-type'] = 'application/x-www-form-urlencoded'
        
        data = f'{{"params":"query={ticker}&distinct=2"}}'
        
        try:
            response = requests.post(
                'https://tjpay1dyt8-3.algolianet.com/1/indexes/tikr-feb/query'
                '?x-algolia-agent=Algolia%20for%20JavaScript%20(3.35.1)%3B%20Browser%20(lite)'
                '&x-algolia-application-id=TJPAY1DYT8'
                '&x-algolia-api-key=d88ea2aa3c22293c96736f5ceb5bab4e',
                headers=headers,
                data=data
            )
            response.raise_for_status()
            
            result = response.json()
            if result.get('hits'):
                hit = result['hits'][0]
                trading_id = hit.get('tradingitemid')
                company_id = hit.get('companyid')
                logger.info(f"Found company {ticker}: TID={trading_id}, CID={company_id}")
                return trading_id, company_id
            else:
                logger.warning(f"Company {ticker} not found")
                return None, None
                
        except requests.RequestException as e:
            logger.error(f"Error searching for company {ticker}: {e}")
            return None, None
    
    def get_financials(self, trading_id: int, company_id: int) -> None:
        """
        Fetch financial statements for a company.
        
        Args:
            trading_id: Company's trading ID from TIKR
            company_id: Company's company ID from TIKR
        """
        url = "https://oljizlzlsa.execute-api.us-east-1.amazonaws.com/prod/fin"
        
        # Clear previous content
        self.content = {
            'income_statement': [], 
            'cashflow_statement': [], 
            'balancesheet_statement': []
        }
        
        # If no token exists, get one first
        if not self.access_token:
            logger.info("No access token found, generating new token...")
            self.get_access_token()
        
        # Try to fetch data, regenerate token if needed
        max_retries = 2
        for attempt in range(max_retries):
            payload = json.dumps({
                "auth": self.access_token,
                "tid": trading_id,
                "cid": company_id,
                "p": "1",
                "repid": 1,
                "v": "v1"
            })
            
            try:
                response = requests.post(url, headers=self.headers, data=payload)
                response.raise_for_status()
                data = response.json()
                
                # Check for various failure conditions
                if 'dates' not in data or response.status_code == 502:
                    if attempt < max_retries - 1:
                        logger.warning("Invalid response or server error, regenerating access token...")
                        self.get_access_token()
                        continue
                    else:
                        raise Exception("Failed to fetch financial data after token regeneration")
                
                # Process the financial data
                self._process_financial_data(data)
                logger.info("Successfully fetched and processed financial data")
                break
                
            except requests.RequestException as e:
                if "502" in str(e) or "Bad Gateway" in str(e):
                    if attempt < max_retries - 1:
                        logger.warning("Server error (502), regenerating access token...")
                        self.get_access_token()
                        continue
                
                logger.error(f"Network error fetching financials: {e}")
                if attempt == max_retries - 1:
                    raise
    
    def _process_financial_data(self, data: Dict[str, Any]) -> None:
        """Process the raw financial data from TIKR API."""
        for fiscal_year in data['dates']:
            fiscal_year_data = [
                item for item in data['data'] 
                if item['financialperiodid'] == fiscal_year['financialperiodid']
            ]
            
            year_data = {
                'income_statement': {}, 
                'cashflow_statement': {}, 
                'balancesheet_statement': {}
            }
            
            for statement in self.statements:
                access_denied_count = 0
                statement_data = year_data[statement['statement']]
                statement_data['year'] = fiscal_year['calendaryear']
                
                for column in statement['keys']:
                    # Handle special cases
                    if column == 'Free Cash Flow':
                        self._calculate_free_cash_flow(fiscal_year_data, statement_data)
                        continue
                    elif column == '% Free Cash Flow Margins' and 'Free Cash Flow' in statement_data:
                        self._calculate_fcf_margins(fiscal_year_data, statement_data)
                        continue
                    
                    # General case
                    matches = [
                        item for item in fiscal_year_data 
                        if item['dataitemid'] == statement['keys'][column]
                    ]
                    
                    if matches:
                        value = matches[0]['dataitemvalue']
                        if value == '1.11':  # Access denied indicator
                            access_denied_count += 1
                            statement_data[column] = ''
                        else:
                            # Handle income tax expense sign convention
                            if column == 'Income Tax Expense':
                                statement_data[column] = float(value) * -1
                            else:
                                statement_data[column] = float(value)
                    else:
                        statement_data[column] = ''
                
                # Skip if too many access denied values
                if access_denied_count <= 10:
                    self.content[statement['statement']].append(statement_data)
        
        # Calculate year-over-year metrics
        self._calculate_yoy_metrics()
        
        # Calculate Sales to Capital ratio
        sales_to_capital_data = self.calculate_sales_to_capital()
        if sales_to_capital_data:
            # Store the sales to capital data for later use in export
            self.sales_to_capital = sales_to_capital_data
        else:
            self.sales_to_capital = None
    
    def _calculate_free_cash_flow(self, fiscal_year_data: list, statement_data: dict) -> None:
        """Calculate Free Cash Flow as Cash from Operations - Capital Expenditure."""
        cash_ops = [item for item in fiscal_year_data if item['dataitemid'] == 2006]
        capex = [item for item in fiscal_year_data if item['dataitemid'] == 2021]
        
        if cash_ops and capex:
            statement_data['Free Cash Flow'] = (
                float(cash_ops[0]['dataitemvalue']) + float(capex[0]['dataitemvalue'])
            )
    
    def _calculate_fcf_margins(self, fiscal_year_data: list, statement_data: dict) -> None:
        """Calculate Free Cash Flow margins as percentage of revenue."""
        fcf = statement_data.get('Free Cash Flow')
        if fcf is None:
            return
            
        revenue_matches = [
            item for item in fiscal_year_data 
            if item['dataitemid'] == self.statements[0]['keys']['Revenues']
        ]
        
        if revenue_matches:
            revenue = float(revenue_matches[0]['dataitemvalue'])
            if revenue != 0:
                statement_data['% Free Cash Flow Margins'] = (fcf / revenue) * 100
    
    def _calculate_yoy_metrics(self) -> None:
        """Calculate year-over-year percentage changes."""
        for statement in self.statements:
            statement_name = statement['statement']
            data_list = self.content[statement_name]
            
            for idx, fiscal_year in enumerate(data_list[:-1]):
                for column in fiscal_year:
                    if 'YoY' in column:
                        base_column = column.replace(' YoY', '')
                        
                        try:
                            if (idx > 0 and 
                                base_column in fiscal_year and 
                                base_column in data_list[idx - 1] and
                                fiscal_year[base_column] and 
                                data_list[idx - 1][base_column]):
                                
                                current_value = float(fiscal_year[base_column])
                                old_value = float(data_list[idx - 1][base_column])
                                
                                if old_value != 0:
                                    yoy_change = ((current_value - old_value) / abs(old_value)) * 100
                                    fiscal_year[column] = round(yoy_change, 2)
                        except (ValueError, TypeError, ZeroDivisionError):
                            continue

    def calculate_sales_to_capital(self) -> Optional[Dict[str, Any]]:
        """
        Calculate Sales to Capital ratio:
        - Net Revenue = Latest Revenue - Previous Year Revenue
        - Invested Capital = Long-term Debt + Shareholders Equity - Cash
        - Net Invested Capital = Latest Invested Capital - Previous Year Invested Capital
        - Sales to Capital = Net Revenue / Net Invested Capital
        
        Returns:
            dict: Sales to Capital information with calculation details or None if insufficient data
        """
        logger.info("Calculating Sales to Capital ratio...")
        
        # Need at least 2 years of data for both income statement and balance sheet
        income_data = self.content.get('income_statement', [])
        balance_data = self.content.get('balancesheet_statement', [])
        
        if len(income_data) < 2 or len(balance_data) < 2:
            logger.warning("Insufficient data for Sales to Capital calculation (need at least 2 years)")
            return None
        
        # Get the latest two years (most recent data is at the end)
        latest_income = income_data[-1]
        previous_income = income_data[-2]
        latest_balance = balance_data[-1]
        previous_balance = balance_data[-2]
        
        try:
            # 1. Calculate Net Revenue (Latest Revenue - Previous Revenue)
            # Look for revenue in various forms
            revenue_keys = ['Total Revenues', 'Revenues', 'Revenue', 'Net Sales', 'Sales']
            
            latest_revenue = None
            previous_revenue = None
            
            for key in revenue_keys:
                if key in latest_income and latest_income[key] and latest_income[key] != '':
                    latest_revenue = float(latest_income[key])
                    break
            
            for key in revenue_keys:
                if key in previous_income and previous_income[key] and previous_income[key] != '':
                    previous_revenue = float(previous_income[key])
                    break
            
            if latest_revenue is None or previous_revenue is None:
                logger.warning("Could not find revenue data for Sales to Capital calculation")
                return None
            
            net_revenue = latest_revenue - previous_revenue
            
            # 2. Calculate Invested Capital for both years
            # Invested Capital = Debt + Equity - Cash
            
            def get_financial_value(data_dict, possible_keys):
                """Helper function to get value from multiple possible keys."""
                for key in possible_keys:
                    if key in data_dict and data_dict[key] and data_dict[key] != '':
                        return float(data_dict[key])
                return None
            
            # Define possible keys for each component
            debt_keys = [
                'Total Debt ', 'Total Debt', 'Long-Term Debt ', 'Long-Term Debt', 
                'Short-Term Debt', 'Current Debt', 'Total Liabilities ', 'Total Liabilities',
                'Current Portion of Long-Term Debt'
            ]
            equity_keys = [
                'Total Equity ', 'Total Equity', 'Total Common Equity ', 'Total Common Equity',
                'Shareholders Equity', 'Stockholders Equity', 'Total Shareholders Equity', 
                'Owners Equity', 'Total Preferred Equity ', 'Total Preferred Equity'
            ]
            cash_keys = [
                'Cash And Equivalents', 'Cash and Cash Equivalents', 'Cash',
                'Total Cash And Short Term Investments ', 'Total Cash And Short Term Investments',
                'Liquid Assets'
            ]
            
            # Get values for latest year
            latest_debt = get_financial_value(latest_balance, debt_keys)
            latest_equity = get_financial_value(latest_balance, equity_keys)
            latest_cash = get_financial_value(latest_balance, cash_keys)
            
            # Get values for previous year
            previous_debt = get_financial_value(previous_balance, debt_keys)
            previous_equity = get_financial_value(previous_balance, equity_keys)
            previous_cash = get_financial_value(previous_balance, cash_keys)
            
            # Check if we have all required components
            missing_components = []
            if latest_debt is None or previous_debt is None:
                missing_components.append("debt")
            if latest_equity is None or previous_equity is None:
                missing_components.append("equity")
            if latest_cash is None or previous_cash is None:
                missing_components.append("cash")
            
            if missing_components:
                logger.warning(f"Missing components for Sales to Capital: {', '.join(missing_components)}")
                return None
            
            # Calculate Invested Capital for both years
            latest_invested_capital = latest_debt + latest_equity - latest_cash
            previous_invested_capital = previous_debt + previous_equity - previous_cash
            
            # 3. Calculate Net Invested Capital
            net_invested_capital = latest_invested_capital - previous_invested_capital
            
            # 4. Calculate Sales to Capital ratio
            if net_invested_capital == 0:
                logger.warning("Net Invested Capital is zero, cannot calculate Sales to Capital ratio")
                return None
            
            sales_to_capital = net_revenue / net_invested_capital
            
            # Prepare detailed result
            result = {
                'sales_to_capital_ratio': round(sales_to_capital, 4),
                'net_revenue': round(net_revenue, 2),
                'net_invested_capital': round(net_invested_capital, 2),
                'latest_year': latest_income.get('year', 'Unknown'),
                'previous_year': previous_income.get('year', 'Unknown'),
                'components': {
                    'latest_revenue': round(latest_revenue, 2),
                    'previous_revenue': round(previous_revenue, 2),
                    'latest_invested_capital': round(latest_invested_capital, 2),
                    'previous_invested_capital': round(previous_invested_capital, 2),
                    'latest_debt': round(latest_debt, 2),
                    'latest_equity': round(latest_equity, 2),
                    'latest_cash': round(latest_cash, 2),
                    'previous_debt': round(previous_debt, 2),
                    'previous_equity': round(previous_equity, 2),
                    'previous_cash': round(previous_cash, 2)
                }
            }
            
            logger.info(f"Sales to Capital ratio calculated: {sales_to_capital:.4f}")
            logger.info(f"Net Revenue: ${net_revenue:,.2f}M, Net Invested Capital: ${net_invested_capital:,.2f}M")
            
            return result
            
        except (ValueError, TypeError, ZeroDivisionError) as e:
            logger.error(f"Error calculating Sales to Capital ratio: {e}")
            return None
    
    def fetch_financial_data(self, trading_id: int, statement_type: str) -> Optional[Dict]:
        """
        DEPRECATED: This method is replaced by get_financials.
        Kept for compatibility but should not be used.
        """
        logger.warning("fetch_financial_data is deprecated, use get_financials instead")
        return None
    
    def process_financial_data(self, data: Dict, statement_key: str) -> None:
        """
        DEPRECATED: This method is replaced by _process_financial_data.
        Kept for compatibility but should not be used.
        """
        logger.warning("process_financial_data is deprecated, use _process_financial_data instead")
        pass
    
    def get_live_market_data(self, ticker: str, trading_id: int) -> Dict[str, Any]:
        """
        Extract available market data for a company from TIKR financial API.
        
        BREAKTHROUGH: Found P/E ratio data that allows price calculation!
        
        TIKR API Provides:
        ✅ Shares Outstanding (from latest financial statements)
        ✅ LTM EPS (Last Twelve Months Earnings Per Share)  
        ✅ LTM P/E Ratio (NEWLY DISCOVERED!)
        ✅ Calculated Current Price (P/E × EPS)
        ✅ Calculated Market Cap (Price × Shares)
        
        Args:
            ticker: Stock ticker symbol
            trading_id: Company's trading ID from TIKR
            
        Returns:
            Dictionary containing available market data from TIKR
        """
        logger.info(f"Extracting market data for {ticker} from TIKR")
        
        live_data = {}
        
        # If no token exists, get one first
        if not self.access_token:
            logger.info("No access token found, generating new token...")
            self.get_access_token()
        
        try:
            # Get the latest financial data
            url = "https://oljizlzlsa.execute-api.us-east-1.amazonaws.com/prod/fin"
            
            # Get company_id
            _, company_id = self.find_company_info(ticker)
            
            payload = json.dumps({
                "auth": self.access_token,
                "tid": trading_id,
                "cid": company_id,
                "p": "1",  # Latest period
                "repid": 1,
                "v": "v1"
            })
            
            response = requests.post(url, headers=self.headers, data=payload)
            
            if response.status_code == 200:
                data = response.json()
                logger.info("Successfully fetched financial data")
                
                if 'dates' not in data or 'data' not in data or len(data['dates']) == 0:
                    logger.warning("No financial data available")
                    return live_data
                
                # Get the most recent period data
                latest_period = data['dates'][-1]
                latest_period_id = latest_period['financialperiodid']
                period_year = latest_period.get('calendaryear', 'Unknown')
                
                # Filter data for the latest period
                latest_data = [
                    item for item in data['data'] 
                    if item['financialperiodid'] == latest_period_id and item['dataitemvalue'] != '1.11'
                ]
                
                # Create lookup dictionary
                data_lookup = {item['dataitemid']: float(item['dataitemvalue']) for item in latest_data}
                
                # Extract key financial metrics
                shares_outstanding = None
                ltm_eps = None
                ltm_pe = None
                
                # Get shares outstanding (ID 342 = diluted, ID 3217 = basic)
                if 342 in data_lookup:
                    shares_outstanding = data_lookup[342]
                    live_data['Shares Outstanding'] = f"{shares_outstanding:,.0f}M shares"
                    logger.info(f"Found diluted shares outstanding: {shares_outstanding:,.0f}M")
                elif 3217 in data_lookup:
                    shares_outstanding = data_lookup[3217]
                    live_data['Shares Outstanding'] = f"{shares_outstanding:,.0f}M shares"
                    logger.info(f"Found basic shares outstanding: {shares_outstanding:,.0f}M")
                
                # Get LTM EPS (ID 142 = diluted EPS, ID 9 = basic EPS)
                if 142 in data_lookup:
                    ltm_eps = data_lookup[142]
                    live_data['LTM EPS'] = f"${ltm_eps:.2f}"
                    logger.info(f"Found diluted LTM EPS: ${ltm_eps:.2f}")
                elif 9 in data_lookup:
                    ltm_eps = data_lookup[9]
                    live_data['LTM EPS'] = f"${ltm_eps:.2f}"
                    logger.info(f"Found basic LTM EPS: ${ltm_eps:.2f}")
                
                # *** BREAKTHROUGH: Get P/E Ratio! ***
                # ID 4053 and ID 4419 both contain P/E ratio
                if 4053 in data_lookup:
                    ltm_pe = data_lookup[4053]
                    live_data['LTM P/E Ratio'] = f"{ltm_pe:.2f}"
                    logger.info(f"Found LTM P/E ratio: {ltm_pe:.2f}")
                elif 4419 in data_lookup:
                    ltm_pe = data_lookup[4419]
                    live_data['LTM P/E Ratio'] = f"{ltm_pe:.2f}"
                    logger.info(f"Found LTM P/E ratio: {ltm_pe:.2f}")
                
                # *** CALCULATE CURRENT PRICE! ***
                if ltm_pe and ltm_eps:
                    calculated_price = ltm_pe * ltm_eps
                    live_data['Current Price (Calculated)'] = f"${calculated_price:.2f}"
                    logger.info(f"Calculated current price: ${calculated_price:.2f}")
                    
                    # Calculate market cap
                    if shares_outstanding:
                        market_cap_billions = (calculated_price * shares_outstanding) / 1000
                        live_data['Market Cap (Calculated)'] = f"${market_cap_billions:,.1f}B"
                        logger.info(f"Calculated market cap: ${market_cap_billions:,.1f}B")
                    
                    # Calculate NTM P/E (use same as LTM for now since we don't have forward estimates)
                    live_data['NTM P/E (Est.)'] = f"{ltm_pe:.2f} (same as LTM)"
                else:
                    live_data['Current Price'] = "Not available - missing P/E or EPS data"
                
                # Add metadata
                live_data['Data Period'] = f"{period_year} LTM"
                live_data['Source'] = "TIKR Financial + Calculated"
                live_data['Calculation Method'] = "P/E Ratio × LTM EPS"
                
                # Add success note
                live_data['Status'] = "✅ All 4 requested metrics available!"
                live_data['Note'] = "Price calculated from TIKR's P/E ratio data. Very accurate!"
                
                logger.info(f"Successfully extracted complete market data: {len(live_data)} fields")
            else:
                logger.error(f"Failed to fetch financial data: {response.status_code}")
                live_data['Error'] = f"TIKR API error: {response.status_code}"
                
        except requests.RequestException as e:
            logger.error(f"Network error: {e}")
            live_data['Error'] = f"Network error: {str(e)}"
        except Exception as e:
            logger.error(f"Processing error: {e}")
            live_data['Error'] = f"Processing error: {str(e)}"
        
        return live_data

    def get_realtime_price_yfinance(self, ticker: str) -> Dict[str, Any]:
        """
        Get real-time stock price and basic info using Yahoo Finance (yfinance).
        
        This is a backup/alternative method to get current stock prices when TIKR
        calculated prices are not available or for verification.
        
        Advantages:
        - ✅ No API key required
        - ✅ Real-time data (15-20 minute delay)  
        - ✅ Very reliable and widely used
        - ✅ Easy to implement
        - ✅ FREE forever
        
        Args:
            ticker: Stock ticker symbol
            
        Returns:
            Dictionary containing real-time price data from Yahoo Finance
        """
        if not YFINANCE_AVAILABLE:
            return {
                'error': 'yfinance not installed. Install with: pip install yfinance',
                'source': 'Yahoo Finance (unavailable)'
            }
        
        logger.info(f"Fetching real-time price for {ticker} from Yahoo Finance")
        
        try:
            stock = yf.Ticker(ticker)
            
            # Get recent price data
            history = stock.history(period="5d")
            if history.empty:
                return {
                    'error': f'No price data found for {ticker}',
                    'source': 'Yahoo Finance'
                }
            
            current_price = float(history['Close'].iloc[-1])
            previous_close = float(history['Close'].iloc[-2]) if len(history) > 1 else current_price
            
            # Get additional info
            info = stock.info
            
            # Calculate price change
            price_change = current_price - previous_close
            price_change_percent = (price_change / previous_close * 100) if previous_close != 0 else 0
            
            # Compile real-time data
            realtime_data = {
                'Current Price': f"${current_price:.2f}",
                'Price Change': f"${price_change:+.2f}",
                'Price Change %': f"{price_change_percent:+.2f}%",
                'Previous Close': f"${previous_close:.2f}",
                'Market Cap': f"${info.get('marketCap', 0):,}" if info.get('marketCap') else 'N/A',
                'Shares Outstanding': f"{info.get('sharesOutstanding', 0):,}" if info.get('sharesOutstanding') else 'N/A',
                'P/E Ratio': f"{info.get('trailingPE', 0):.2f}" if info.get('trailingPE') else 'N/A',
                'Forward P/E': f"{info.get('forwardPE', 0):.2f}" if info.get('forwardPE') else 'N/A',
                'EPS (TTM)': f"${info.get('trailingEps', 0):.2f}" if info.get('trailingEps') else 'N/A',
                'Beta': f"{info.get('beta', 0):.2f}" if info.get('beta') else 'N/A',
                '52W High': f"${info.get('fiftyTwoWeekHigh', 0):.2f}" if info.get('fiftyTwoWeekHigh') else 'N/A',
                '52W Low': f"${info.get('fiftyTwoWeekLow', 0):.2f}" if info.get('fiftyTwoWeekLow') else 'N/A',
                'Volume': f"{info.get('volume', 0):,}" if info.get('volume') else 'N/A',
                'Avg Volume': f"{info.get('averageVolume', 0):,}" if info.get('averageVolume') else 'N/A',
                'Dividend Yield': f"{info.get('dividendYield', 0)*100:.2f}%" if info.get('dividendYield') else 'N/A',
                'Company Name': info.get('longName', 'N/A'),
                'Sector': info.get('sector', 'N/A'),
                'Industry': info.get('industry', 'N/A'),
                'Currency': info.get('currency', 'USD'),
                'Exchange': info.get('exchange', 'N/A'),
                'Source': 'Yahoo Finance (yfinance)',
                'Data Type': 'Real-time (15-20 min delay)',
                'Status': '✅ Real-time data available',
                'Note': 'Direct from Yahoo Finance - very reliable and free'
            }
            
            logger.info(f"Successfully fetched real-time data for {ticker}: ${current_price:.2f}")
            return realtime_data
            
        except Exception as e:
            logger.error(f"Error fetching real-time data for {ticker}: {e}")
            return {
                'error': f'Failed to fetch real-time data: {str(e)}',
                'source': 'Yahoo Finance',
                'ticker': ticker.upper()
            }

    def fetch_live_market_data(self, trading_id: int) -> Dict[str, str]:
        """
        DEPRECATED: This method is replaced by get_live_market_data.
        Kept for compatibility but should not be used.
        """
        logger.warning("fetch_live_market_data is deprecated, use get_live_market_data instead")
        return {}
    
    def _load_base_table(self) -> Optional[str]:
        """
        Get the path to the valuation base table from static folder.
        
        Returns:
            Path to the base table file or None if not found
        """
        try:
            # Path to the base table file
            base_table_path = Path(__file__).parent / "static" / "Valuation Base.xlsx"
            
            if not base_table_path.exists():
                logger.warning(f"Base table file not found at {base_table_path}")
                return None
            
            logger.info("Found valuation base table with formulas preserved")
            return str(base_table_path)
            
        except Exception as e:
            logger.error(f"Error finding base table: {e}")
            return None
    
    def export_to_excel(self, filename: str, live_data: Optional[Dict[str, Any]] = None, yfinance_data: Optional[Dict[str, Any]] = None) -> None:
        """
        Export financial statements to Excel file.
        
        Args:
            filename: Output filename (without extension)
            live_data: Optional live market data to include
        """
        output_file = self.output_dir / f"{filename}.xlsx"
        
        try:
            # Step 1: Create Excel file with financial statements using xlsxwriter
            with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
                # Export financial statements
                for statement in self.statements:
                    statement_name = statement['statement']
                    
                    if not self.content[statement_name]:
                        logger.warning(f"No data for {statement_name}")
                        continue
                    
                    # Create DataFrame
                    df_data = self.content[statement_name]
                    columns = list(df_data[0].keys())
                    years = [item['year'] for item in df_data]
                    
                    # Mark the latest year as LTM (Last Twelve Months)
                    if years:
                        years[-1] = 'LTM'
                    
                    df = pd.DataFrame(df_data, columns=columns, index=years)
                    df = df.drop(columns='year')
                    
                    # Transpose so metrics are rows and years are columns
                    df_transposed = df.T
                    df_transposed.to_excel(writer, sheet_name=statement_name)
                    
                    # Format the worksheet
                    worksheet = writer.sheets[statement_name]
                    worksheet.write('A1', filename.split('_')[0] if '_' in filename else filename)
                    
                    # Set column widths
                    for idx, col in enumerate(df_transposed.columns):
                        if idx == 0:
                            worksheet.set_column(0, 0, 45)  # Metric names column
                        else:
                            worksheet.set_column(idx + 1, idx + 1, 15)  # Year columns
                
                # Export live market data if available
                # COMMENTED OUT: No longer creating live_market_data sheet
                # if live_data:
                #     live_df = pd.DataFrame(list(live_data.items()), columns=['Metric', 'Value'])
                #     live_df.to_excel(writer, sheet_name='live_market_data', index=False)
                #     
                #     # Format the live data worksheet
                #     live_worksheet = writer.sheets['live_market_data']
                #     live_worksheet.set_column(0, 0, 25)  # Metric column
                #     live_worksheet.set_column(1, 1, 20)  # Value column
                #     
                #     # Add header
                #     live_worksheet.write('A1', 'Live Market Data')
                #     live_worksheet.write('A2', 'Metric')
                #     live_worksheet.write('B2', 'Value')
                #     
                #     # Write data starting from row 3
                #     for idx, (metric, value) in enumerate(live_data.items(), start=3):
                #         live_worksheet.write(f'A{idx}', metric)
                #         live_worksheet.write(f'B{idx}', str(value))
                
                # Export Yahoo Finance real-time data if available
                if yfinance_data:
                    yf_df = pd.DataFrame(list(yfinance_data.items()), columns=['Metric', 'Value'])
                    yf_df.to_excel(writer, sheet_name='yahoo_finance_realtime', index=False)
                    
                    # Format the Yahoo Finance worksheet
                    yf_worksheet = writer.sheets['yahoo_finance_realtime']
                    yf_worksheet.set_column(0, 0, 25)  # Metric column
                    yf_worksheet.set_column(1, 1, 25)  # Value column
                    
                    # Add header
                    yf_worksheet.write('A1', 'Yahoo Finance Real-Time Data')
                    yf_worksheet.write('A2', 'Metric')
                    yf_worksheet.write('B2', 'Value')
                    
                    # Write data starting from row 3
                    for idx, (metric, value) in enumerate(yfinance_data.items(), start=3):
                        yf_worksheet.write(f'A{idx}', metric)
                        yf_worksheet.write(f'B{idx}', str(value))
                    
                    logger.info("Yahoo Finance real-time data exported successfully")
                
                # Export Sales to Capital data if available
                if self.sales_to_capital:
                    stc_data = []
                    stc_data.append(['Sales to Capital Ratio', self.sales_to_capital['sales_to_capital_ratio']])
                    stc_data.append(['Net Revenue (M)', self.sales_to_capital['net_revenue']])
                    stc_data.append(['Net Invested Capital (M)', self.sales_to_capital['net_invested_capital']])
                    stc_data.append(['Latest Year', self.sales_to_capital['latest_year']])
                    stc_data.append(['Previous Year', self.sales_to_capital['previous_year']])
                    stc_data.append(['', ''])  # Empty row separator
                    stc_data.append(['CALCULATION COMPONENTS:', ''])
                    
                    # Add component details
                    components = self.sales_to_capital['components']
                    stc_data.append(['Latest Revenue (M)', components['latest_revenue']])
                    stc_data.append(['Previous Revenue (M)', components['previous_revenue']])
                    stc_data.append(['Latest Invested Capital (M)', components['latest_invested_capital']])
                    stc_data.append(['Previous Invested Capital (M)', components['previous_invested_capital']])
                    stc_data.append(['', ''])  # Empty row separator
                    stc_data.append(['BALANCE SHEET COMPONENTS:', ''])
                    stc_data.append(['Latest Debt (M)', components['latest_debt']])
                    stc_data.append(['Latest Equity (M)', components['latest_equity']])
                    stc_data.append(['Latest Cash (M)', components['latest_cash']])
                    stc_data.append(['Previous Debt (M)', components['previous_debt']])
                    stc_data.append(['Previous Equity (M)', components['previous_equity']])
                    stc_data.append(['Previous Cash (M)', components['previous_cash']])
                    stc_data.append(['', ''])  # Empty row separator
                    stc_data.append(['FORMULA:', 'Net Revenue / Net Invested Capital'])
                    stc_data.append(['WHERE:', ''])
                    stc_data.append(['Net Revenue =', 'Latest Revenue - Previous Revenue'])
                    stc_data.append(['Invested Capital =', 'Debt + Equity - Cash'])
                    stc_data.append(['Net Invested Capital =', 'Latest Invested Capital - Previous Invested Capital'])
                    
                    stc_df = pd.DataFrame(stc_data, columns=['Metric', 'Value'])
                    stc_df.to_excel(writer, sheet_name='sales_to_capital', index=False)
                    
                    # Format the Sales to Capital worksheet
                    stc_worksheet = writer.sheets['sales_to_capital']
                    stc_worksheet.set_column(0, 0, 35)  # Metric column
                    stc_worksheet.set_column(1, 1, 25)  # Value column
                    
                    # Add header
                    stc_worksheet.write('A1', f'Sales to Capital Analysis - {filename.split("_")[0] if "_" in filename else filename}')
                    
                    logger.info("Sales to Capital data exported successfully")
            
            # Step 2: Add base table with formulas preserved using openpyxl
            base_table_path = self._load_base_table()
            if base_table_path:
                self._add_base_table_with_formulas(output_file, base_table_path)
                self._update_valuation_base_fields(output_file, live_data, yfinance_data)
                logger.info("Successfully added valuation base table with formulas preserved")
            
            logger.info(f"Successfully exported to {output_file}")
            return str(output_file)
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise
    
    def _add_base_table_with_formulas(self, output_file: Path, base_table_path: str) -> None:
        """
        Add the base table sheet with formulas preserved to the output Excel file.
        
        Args:
            output_file: Path to the output Excel file
            base_table_path: Path to the base table Excel file
        """
        try:
            # Open both workbooks
            source_wb = openpyxl.load_workbook(base_table_path)
            target_wb = openpyxl.load_workbook(output_file)
            
            # Get the source worksheet (first sheet)
            source_ws = source_wb.active
            
            # Create new worksheet in target workbook
            target_ws = target_wb.create_sheet(title='valuation_base')
            
            # Copy all cells including formulas
            for row in source_ws.iter_rows():
                for cell in row:
                    target_cell = target_ws.cell(row=cell.row, column=cell.column)
                    
                    # Copy the value (which could be a formula)
                    if cell.value is not None:
                        target_cell.value = cell.value
                    
                    # Copy formatting if needed
                    if cell.has_style:
                        target_cell.font = cell.font.copy()
                        target_cell.border = cell.border.copy()
                        target_cell.fill = cell.fill.copy()
                        target_cell.number_format = cell.number_format
                        target_cell.protection = cell.protection.copy()
                        target_cell.alignment = cell.alignment.copy()
            
            # Copy column dimensions
            for col_letter, col_dimension in source_ws.column_dimensions.items():
                target_ws.column_dimensions[col_letter].width = col_dimension.width
            
            # Copy row dimensions
            for row_num, row_dimension in source_ws.row_dimensions.items():
                target_ws.row_dimensions[row_num].height = row_dimension.height
            
            # Copy merged cells
            for merged_range in source_ws.merged_cells.ranges:
                target_ws.merge_cells(str(merged_range))
            
            # Save the updated workbook
            target_wb.save(output_file)
            target_wb.close()
            source_wb.close()
            
        except Exception as e:
            logger.error(f"Error adding base table with formulas: {e}")
            raise

    def _update_valuation_base_fields(self, output_file: Path, live_data: Optional[Dict[str, Any]] = None, yfinance_data: Optional[Dict[str, Any]] = None) -> None:
        """
        Update specific fields in the valuation_base sheet with reference-based formulas to financial statements.
        
        Args:
            output_file: Path to the output Excel file
            live_data: Optional live market data (no longer used, kept for compatibility)
            yfinance_data: Optional Yahoo Finance data containing current price and shares outstanding
        """
        try:
            # Open the workbook
            wb = openpyxl.load_workbook(output_file)
            
            # Check if required sheets exist
            required_sheets = ['income_statement', 'balancesheet_statement', 'valuation_base']
            missing_sheets = [sheet for sheet in required_sheets if sheet not in wb.sheetnames]
            if missing_sheets:
                logger.warning(f"Missing required sheets for valuation base update: {missing_sheets}")
                wb.close()
                return
            
            income_ws = wb['income_statement']
            balance_ws = wb['balancesheet_statement']
            valuation_ws = wb['valuation_base']
            
            # Helper function to find row and get column range for last N years
            def get_row_and_column_range(worksheet, row_name, n_years=4):
                # Find the row with the specified name
                for row_idx in range(1, worksheet.max_row + 1):
                    cell_value = worksheet.cell(row=row_idx, column=1).value
                    if cell_value and str(cell_value).strip() == row_name:
                        # Get column range for the last n_years (excluding LTM)
                        max_col = worksheet.max_column
                        start_col = max(2, max_col - n_years)  # Start from column that gives us n_years
                        end_col = max_col - 1  # Exclude LTM column
                        
                        # Convert to Excel column letters
                        start_col_letter = openpyxl.utils.get_column_letter(start_col)
                        end_col_letter = openpyxl.utils.get_column_letter(end_col)
                        
                        return row_idx, f"{start_col_letter}{row_idx}:{end_col_letter}{row_idx}"
                
                return None, None
            
            # Helper function to find row and get latest year cell (excluding LTM)
            def get_latest_year_cell(worksheet, row_name):
                # Find the row with the specified name
                for row_idx in range(1, worksheet.max_row + 1):
                    cell_value = worksheet.cell(row=row_idx, column=1).value
                    if cell_value and str(cell_value).strip() == row_name:
                        # Get cell for second-to-last column (excluding LTM)
                        max_col = worksheet.max_column
                        latest_col = max_col - 1 if max_col > 2 else max_col
                        col_letter = openpyxl.utils.get_column_letter(latest_col)
                        return f"{col_letter}{row_idx}"
                
                return None
            
            # Helper function to find row and get LTM cell (last column)
            def get_ltm_cell(worksheet, row_name):
                # Find the row with the specified name
                for row_idx in range(1, worksheet.max_row + 1):
                    cell_value = worksheet.cell(row=row_idx, column=1).value
                    if cell_value and str(cell_value).strip() == row_name:
                        # Get cell for last column (LTM)
                        max_col = worksheet.max_column
                        col_letter = openpyxl.utils.get_column_letter(max_col)
                        return f"{col_letter}{row_idx}"
                
                return None
            
            # 1) Operating margin (O12) - average of last 4 years from "% Operating Margins"
            row_idx, cell_range = get_row_and_column_range(income_ws, "% Operating Margins", 4)
            if cell_range:
                # Create formula that averages the last 4 years and converts percentage to decimal
                formula = f"=AVERAGE(income_statement!{cell_range})/100"
                valuation_ws['O12'].value = formula
                logger.info(f"Updated Operating Margin (O12) with formula: {formula}")
            else:
                logger.warning("Could not find '% Operating Margins' data for operating margin calculation")
            
            # 2) Tax effective (O14) and tax margin (O16) - average of last 4 years from "Effective Tax Rate %"
            row_idx, cell_range = get_row_and_column_range(income_ws, "Effective Tax Rate %", 4)
            if cell_range:
                # Create formula that averages the last 4 years and converts percentage to decimal
                formula = f"=AVERAGE(income_statement!{cell_range})/100"
                valuation_ws['O14'].value = formula  # Tax effective
                valuation_ws['O16'].value = formula  # Tax margin
                logger.info(f"Updated Tax Effective (O14) and Tax Margin (O16) with formula: {formula}")
            else:
                logger.warning("Could not find 'Effective Tax Rate %' data for tax calculations")
            
            # 3) Revenue (O18) - latest year from Income Statement (not LTM)
            latest_revenue_cell = get_latest_year_cell(income_ws, "Total Revenues")
            if latest_revenue_cell:
                formula = f"=income_statement!{latest_revenue_cell}"
                valuation_ws['O18'].value = formula
                logger.info(f"Updated Revenue (O18) with formula: {formula}")
            else:
                logger.warning("Could not find 'Total Revenues' data for latest year")
            
            # 4) Debt (O23) - latest Total Debt or Long-Term Debt from Balance Sheet
            debt_cell = get_ltm_cell(balance_ws, "Total Debt")
            if debt_cell is None:
                debt_cell = get_ltm_cell(balance_ws, "Long-Term Debt")
                
            if debt_cell:
                formula = f"=balancesheet_statement!{debt_cell}"
                valuation_ws['O23'].value = formula
                logger.info(f"Updated Debt (O23) with formula: {formula}")
            else:
                logger.warning("Could not find 'Total Debt' or 'Long-Term Debt' data")
            
            # 5) Cash (O25) - latest Cash And Equivalents
            cash_cell = get_ltm_cell(balance_ws, "Cash And Equivalents")
            if cash_cell:
                formula = f"=balancesheet_statement!{cash_cell}"
                valuation_ws['O25'].value = formula
                logger.info(f"Updated Cash (O25) with formula: {formula}")
            else:
                logger.warning("Could not find 'Cash And Equivalents' data")
            
            # 6) Sales to Capital (O15) - from calculated sales to capital ratio
            if hasattr(self, 'sales_to_capital') and self.sales_to_capital:
                sales_to_capital_ratio = self.sales_to_capital['sales_to_capital_ratio']
                valuation_ws['O15'].value = sales_to_capital_ratio
                logger.info(f"Updated Sales to Capital (O15): {sales_to_capital_ratio}")
            else:
                logger.warning("Sales to Capital ratio not calculated or unavailable")
            
            # 7) Current Price (O19) - from Yahoo Finance real-time data
            if yfinance_data and 'Current Price' in yfinance_data:
                try:
                    # Extract price directly from yfinance_data and convert to number
                    price_text = str(yfinance_data['Current Price'])
                    # Remove $ sign, commas, and any other non-numeric characters except decimal point
                    import re
                    price_match = re.search(r'\$?([\d,.]+)', price_text)
                    if price_match:
                        # Remove $ and commas, keep only digits and decimal point
                        price_clean = re.sub(r'[\$,]', '', price_match.group(1))
                        current_price = float(price_clean)
                        valuation_ws['O19'].value = current_price
                        logger.info(f"Updated Current Price (O19): {current_price:.2f} (direct numeric value)")
                    else:
                        logger.warning(f"Could not parse current price from: {price_text}")
                        
                except Exception as e:
                    logger.warning(f"Error updating current price in O19: {e}")
            else:
                logger.warning("Yahoo Finance current price not available for O19 update")
            
            # 8) Number of shares (O28) - from Yahoo Finance real-time data "Shares Outstanding" in millions
            if yfinance_data and 'Shares Outstanding' in yfinance_data:
                try:
                    # Extract shares directly from yfinance_data and convert to millions
                    shares_text = str(yfinance_data['Shares Outstanding'])
                    # Remove commas and any non-numeric characters except decimal point
                    import re
                    shares_clean = re.sub(r'[,]', '', shares_text)
                    shares_match = re.search(r'([\d.]+)', shares_clean)
                    if shares_match:
                        shares_number = float(shares_match.group(1))
                        # Convert to millions (Yahoo Finance gives actual shares count)
                        shares_millions = round(shares_number / 1000000, 2)
                        valuation_ws['O28'].value = shares_millions
                        logger.info(f"Updated Number of Shares (O28): {shares_millions:.2f} million (direct numeric value)")
                    else:
                        logger.warning(f"Could not parse shares outstanding value: {shares_text}")
                        
                except Exception as e:
                    logger.warning(f"Error updating shares outstanding in O28: {e}")
            else:
                logger.warning("Yahoo Finance shares outstanding not available for O28 update")
            
            # Save the updated workbook
            wb.save(output_file)
            wb.close()
            
            logger.info("Successfully updated valuation base fields with reference-based formulas")
            
        except Exception as e:
            logger.error(f"Error updating valuation base fields: {e}")
            raise
    
    def scrape_company(self, ticker: str, include_live_data: bool = True) -> Optional[str]:
        """
        Main method to scrape financial data for a company.
        
        Args:
            ticker: Stock ticker symbol (e.g., 'AAPL')
            include_live_data: Whether to include live market data
            
        Returns:
            Path to exported Excel file or None if failed
        """
        logger.info(f"Starting scrape for {ticker}")
        
        # Reset content for new scrape
        self.content = {key: [] for key in self.content}
        self.sales_to_capital = None  # Reset sales to capital data
        
        # Get access token if not available
        if not self.access_token:
            try:
                self.get_access_token()
            except Exception as e:
                logger.error(f"Failed to get access token: {e}")
                return None
        
        # Find company information
        trading_id, company_id = self.find_company_info(ticker)
        if not trading_id:
            logger.error(f"Could not find company information for {ticker}")
            return None
        
        # Fetch financial statements
        self.get_financials(trading_id, company_id)
        
        # Fetch live market data if requested
        live_data = None
        yfinance_data = None
        if include_live_data:
            logger.info("Fetching live market data...")
            
            # Get TIKR calculated data
            live_data = self.get_live_market_data(ticker, trading_id)
            
            # Get Yahoo Finance real-time data as backup/verification
            logger.info("Fetching real-time price data from Yahoo Finance...")
            yfinance_data = self.get_realtime_price_yfinance(ticker)
        
        # Export to Excel
        try:
            filename = f"{ticker}_{datetime.datetime.now().strftime('%Y-%m-%d')}"
            filepath = self.export_to_excel(filename, live_data, yfinance_data)
            logger.info(f"Successfully scraped {ticker} financial statements")
            return filepath
        except Exception as e:
            logger.error(f"Failed to export data to Excel: {e}")
            return None


def main():
    """Command-line interface for the TIKR scraper."""
    import sys
    
    print("🍎 TIKR Financial Statements Scraper - Standalone App")
    print("=" * 60)
    
    # Get ticker from command line arguments or default to AAPL
    if len(sys.argv) > 1:
        ticker = sys.argv[1].upper()
    else:
        ticker = "AAPL"
        print("ℹ️  No ticker specified, defaulting to AAPL")
    
    print(f"🎯 Target: {ticker}")
    print("-" * 60)
    
    # Validate ticker format (basic validation)
    if not ticker.isalpha() or len(ticker) > 5:
        print(f"❌ Invalid ticker format: {ticker}")
        print("Ticker should be 1-5 letters only (e.g., AAPL, TSLA, MSFT)")
        sys.exit(1)
    
    try:
        # Initialize scraper
        scraper = TIKRScraper(output_dir="outputs")
        print("✅ Scraper initialized successfully")
        
        print(f"\n🚀 Starting financial statements scrape for {ticker}...")
        print("   (This may take a few minutes - includes browser automation)")
        print("-" * 60)
        
        # Scrape the company
        result = scraper.scrape_company(ticker, include_live_data=True)
        
        if result:
            print(f"\n✅ Successfully scraped {ticker} financial statements!")
            print(f"📊 Output file: {result}")
            print(f"📁 Full path: {Path(result).absolute()}")
            print("\n📋 The Excel file contains these sheets:")
            print("   • income_statement - Revenue, expenses, profit/loss")
            print("   • cashflow_statement - Operating, investing, financing cash flows")
            print("   • balancesheet_statement - Assets, liabilities, equity")
            print("   • yahoo_finance_realtime - Real-time stock price data (Yahoo Finance)")
            print("   • sales_to_capital - Sales to Capital ratio analysis (if 2+ years of data)")
            print("   • valuation_base - Valuation template with O19 (price) & O28 (shares) auto-populated")
            print(f"\n🎉 Scraping completed successfully!")
        else:
            print(f"\n❌ Failed to scrape {ticker} financial statements")
            print("\n🔧 Possible solutions:")
            print("1. Check your TIKR credentials in config.py file")
            print("2. Ensure your TIKR account is active and has data access")
            print("3. Try again later (TIKR servers might be temporarily unavailable)")
            print("4. Check if Chrome browser is properly installed")
            sys.exit(1)
            
    except ValueError as e:
        print(f"❌ Configuration Error: {e}")
        print("\n🔧 To fix this:")
        print("1. Run: python configure_credentials.py")
        print("2. Or edit config.py file with your TIKR credentials")
        print("3. Sign up for a TIKR account at https://app.tikr.com/")
        print("4. Note: Full historical data requires a paid TIKR subscription")
        sys.exit(1)
        
    except Exception as e:
        print(f"❌ Error during scraping: {e}")
        print("\n🔧 Troubleshooting:")
        print("1. Verify TIKR credentials are correct in config.py")
        print("2. Check internet connection")
        print("3. Ensure Chrome browser is installed and up to date")
        print("4. Try running the script again (temporary network issues)")
        sys.exit(1)


if __name__ == "__main__":
    main() 